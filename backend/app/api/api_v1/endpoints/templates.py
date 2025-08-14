# File: app/api/api_v1/endpoints/templates.py
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import Optional
import os

from app.db.database import get_db
from app.api.api_v1.endpoints.auth import get_current_user
from app.models.user import User
from app.services.template_service import TemplateService

router = APIRouter()


@router.get("/credit-card/download")
async def download_credit_card_template(
        include_sample: bool = Query(True, description="Include sample data in template"),
        current_user: User = Depends(get_current_user)
):
    """Download Excel template for Credit Card defaulter accounts"""

    template_service = TemplateService()

    try:
        template_path = template_service.generate_template("CREDIT_CARD", include_sample=include_sample)

        if not os.path.exists(template_path):
            raise HTTPException(status_code=500, detail="Template generation failed")

        filename = "RECOVA_Credit_Card_Import_Template.xlsx"
        return FileResponse(
            path=template_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate template: {str(e)}")


@router.get("/loan/download")
async def download_loan_template(
        include_sample: bool = Query(True, description="Include sample data in template"),
        current_user: User = Depends(get_current_user)
):
    """Download Excel template for Loan defaulter accounts"""

    template_service = TemplateService()

    try:
        template_path = template_service.generate_template("LOAN", include_sample=include_sample)

        if not os.path.exists(template_path):
            raise HTTPException(status_code=500, detail="Template generation failed")

        filename = "RECOVA_Loan_Import_Template.xlsx"
        return FileResponse(
            path=template_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate template: {str(e)}")


@router.get("/leasing/download")
async def download_leasing_template(
        include_sample: bool = Query(True, description="Include sample data in template"),
        current_user: User = Depends(get_current_user)
):
    """Download Excel template for Leasing/Vehicle Finance defaulter accounts"""

    template_service = TemplateService()

    try:
        template_path = template_service.generate_template("LEASING", include_sample=include_sample)

        if not os.path.exists(template_path):
            raise HTTPException(status_code=500, detail="Template generation failed")

        filename = "RECOVA_Leasing_Import_Template.xlsx"
        return FileResponse(
            path=template_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate template: {str(e)}")


@router.get("/list")
async def list_available_templates(current_user: User = Depends(get_current_user)):
    """List all available import templates with their details"""

    template_service = TemplateService()
    template_info = template_service.get_template_info()

    return {
        "templates": [
            {
                "operation_type": "CREDIT_CARD",
                "name": "Credit Card Accounts",
                "description": template_info["CREDIT_CARD"]["description"],
                "field_count": template_info["CREDIT_CARD"]["field_count"],
                "download_url": "/api/v1/templates/credit-card/download",
                "sample_available": template_info["CREDIT_CARD"]["sample_provided"]
            },
            {
                "operation_type": "LOAN",
                "name": "Loan Accounts",
                "description": template_info["LOAN"]["description"],
                "field_count": template_info["LOAN"]["field_count"],
                "download_url": "/api/v1/templates/loan/download",
                "sample_available": template_info["LOAN"]["sample_provided"]
            },
            {
                "operation_type": "LEASING",
                "name": "Leasing/Vehicle Finance Accounts",
                "description": template_info["LEASING"]["description"],
                "field_count": template_info["LEASING"]["field_count"],
                "download_url": "/api/v1/templates/leasing/download",
                "sample_available": template_info["LEASING"]["sample_provided"]
            }
        ],
        "instructions": {
            "usage": "Download the appropriate template based on the type of defaulter accounts you want to import",
            "sample_data": "All templates include sample data by default. Use ?include_sample=false to get blank templates",
            "import_process": "After filling the template, use the /api/v1/imports/upload endpoint to import the data"
        }
    }


@router.get("/payment/download")
async def download_payment_template(
        include_sample: bool = Query(True, description="Include sample data in template"),
        current_user: User = Depends(get_current_user)
):
    """Download Excel template for daily payment imports"""

    # Simple payment template
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment_Import_Template"

    # Header style
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Payment template headers with bank information
    headers = [
        "bank_name", "operation_name",  # NEW: Bank/Client information
        "payment_date", "contract_number", "account_number", "customer_nic",
        "payment_amount", "payment_type", "receipt_number", "bank_reference_number",
        "payment_method", "branch_name", "payment_remarks", "current_total_arrears"
    ]

    # Add title
    ws.merge_cells('A1:F1')
    ws['A1'] = "RECOVA - Payment Import Template"
    ws['A1'].font = Font(name="Arial", size=14, bold=True, color="366092")
    ws['A1'].alignment = Alignment(horizontal="center")

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header.replace("_", " ").title()
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add sample data if requested
    if include_sample:
        sample_data = [
            ["Commercial Bank", "Payments", "2025-08-13", "CC001234567", "ACC123456", "123456789V", 25000.00, "PARTIAL",
             "RCP001", "BNK123456", "Bank Transfer", "Colombo Central", "Monthly payment", 240000.00],
            ["Sampath Bank", "Payments", "2025-08-13", "LN987654321", "OP123456789", "234567890V", 15000.00,
             "INSTALLMENT", "RCP002", "BNK123457", "Cash", "Kandy Main", "Regular installment", 785000.00],
            ["Central Finance", "Payments", "2025-08-13", "LS456789012", "", "345678901V", 18000.00, "RENTAL", "RCP003",
             "BNK123458", "Cheque", "Galle Fort", "Monthly rental", 72000.00]
        ]

        for row_idx, row_data in enumerate(sample_data, 4):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust columns - Fixed version
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0

        # Check header length
        header_cell = ws.cell(row=3, column=col_num)
        if header_cell.value:
            max_length = max(max_length, len(str(header_cell.value)))

        # Check sample data length if exists
        if include_sample:
            for row_num in range(4, 7):  # Sample data rows
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        # Set width
        adjusted_width = min(max(max_length + 2, 12), 40)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save template
    template_dir = "templates"
    os.makedirs(template_dir, exist_ok=True)
    filename = "RECOVA_Payment_Import_Template.xlsx"
    filepath = os.path.join(template_dir, filename)
    wb.save(filepath)

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )