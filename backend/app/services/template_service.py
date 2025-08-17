# backend/app/services/template_service.py - COMPLETE VERSION WITH ALL FIELDS
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date


class TemplateService:
    def __init__(self):
        self.header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def get_all_customer_fields(self):
        """Get all customer fields from your Customer model"""
        return [
            # Client and Contract Information
            "client_code",
            "contract_number",

            # Personal Information
            "client_name",
            "nic",
            "home_address",

            # Contact Details
            "customer_contact_number_1",
            "customer_contact_number_2",
            "customer_contact_number_3",

            # Financial Information
            "granted_amount",
            "facility_granted_date",
            "facility_end_date",
            "monthly_rental_payment_with_vat",
            "last_payment_date",
            "last_payment_amount",
            "due_date",

            # Employment Information
            "designation",
            "work_place_name",
            "work_place_address",
            "work_place_contact_number_1",
            "work_place_contact_number_2",
            "work_place_contact_number_3",

            # Guarantor 1 Information
            "guarantor_1_name",
            "guarantor_1_address",
            "guarantor_1_nic",
            "guarantor_1_contact_number_1",
            "guarantor_1_contact_number_2",

            # Guarantor 2 Information
            "guarantor_2_name",
            "guarantor_2_address",
            "guarantor_2_nic",
            "guarantor_2_contact_number_1",
            "guarantor_2_contact_number_2",

            # Location Information
            "zone",
            "region",
            "branch_name",
            "district_name",
            "postal_town",

            # Status Information
            "days_in_arrears",
            "details",
            "payment_assumption"
        ]

    def get_credit_card_specific_fields(self):
        """Credit card specific fields"""
        return [
            "card_product_type",
            "card_number",
            "account_number",
            "credit_limit",
            "capital_balance",
            "interest_over_due_balance",
            "total_outstanding_amount",
            "minimum_due_amount"
        ]

    def get_loan_specific_fields(self):
        """Loan specific fields"""
        return [
            "loan_type",
            "loan_amount",
            "installment_amount",
            "outstanding_balance",
            "interest_rate",
            "loan_term_months",
            "remaining_installments"
        ]

    def get_leasing_specific_fields(self):
        """Leasing specific fields"""
        return [
            "vehicle_number",
            "vehicle_make",
            "vehicle_model",
            "vehicle_year",
            "lease_amount",
            "lease_term_months",
            "monthly_rental",
            "outstanding_balance",
            "residual_value"
        ]

    def get_sample_data_credit_card(self):
        """Sample data for credit card template"""
        return [
            [
                "SAMPBANK", "CC001234567", "John Doe", "123456789V", "123 Main St, Colombo 03",
                "0771234567", "0112345678", "0774567890", 500000.00, "2020-01-15", "2025-01-15",
                25000.00, "2025-07-15", 20000.00, "2025-08-15", "Manager", "ABC Company Ltd",
                "456 Business St, Colombo 02", "0112223333", "0114445555", "0776667777",
                "Jane Doe", "789 Family St, Dehiwala", "234567890V", "0711111111", "0112222222",
                "Bob Smith", "321 Friend St, Nugegoda", "345678901V", "0722222222", "0113333333",
                "Western", "Colombo", "Colombo Central", "Colombo", "Colombo", 15,
                "Regular customer with good payment history", "Monthly payment expected",
                "Visa", "****1234", "ACC12345678", 500000.00, 125000.00, 15000.00, 140000.00, 25000.00
            ],
            [
                "COMBANK", "CC987654321", "Jane Smith", "234567890V", "456 Second St, Kandy",
                "0759876543", "0812345678", "", 300000.00, "2021-03-20", "2026-03-20",
                15000.00, "2025-06-20", 12000.00, "2025-08-20", "Executive", "XYZ Corp",
                "789 Office Ave, Kandy", "0812223344", "", "", "Alice Johnson", "654 Sister St, Peradeniya",
                "456789012V", "0773333333", "", "", "", "", "", "", "", "Central", "Kandy",
                "Kandy Main", "Kandy", "Kandy", 8, "Good customer", "Regular payments",
                "MasterCard", "****5678", "ACC87654321", 300000.00, 85000.00, 8000.00, 93000.00, 15000.00
            ]
        ]

    def get_sample_data_loan(self):
        """Sample data for loan template"""
        return [
            [
                "SAMPBANK", "LN001234567", "Alice Johnson", "345678901V", "789 Third St, Galle",
                "0771111111", "0912222222", "", 1000000.00, "2022-01-10", "2027-01-10",
                25000.00, "2025-07-10", 22000.00, "2025-08-10", "Director", "DEF Industries",
                "111 Industrial Rd, Galle", "0912345678", "0919876543", "", "Mike Johnson",
                "222 Brother St, Matara", "567890123V", "0774444444", "0913333333", "Sarah Wilson",
                "333 Sister St, Hikkaduwa", "678901234V", "0775555555", "", "Southern", "Galle",
                "Galle Fort", "Galle", "Galle", 25, "Housing loan customer", "Monthly installments",
                "Housing Loan", 1000000.00, 25000.00, 750000.00, 8.5, 60, 35
            ],
            [
                "SEYLAN", "LN987654321", "Bob Wilson", "456789012V", "321 Fourth St, Matara",
                "0772222222", "0412345678", "0778888888", 2500000.00, "2023-06-15", "2028-06-15",
                45000.00, "2025-05-15", 40000.00, "2025-08-15", "Manager", "GHI Trading",
                "444 Trade St, Matara", "0412223344", "", "", "Lisa Wilson", "555 Wife St, Weligama",
                "789012345V", "0776666666", "", "", "", "", "", "", "", "Southern", "Matara",
                "Matara Main", "Matara", "Matara", 12, "Business loan", "Regular payments",
                "Business Loan", 2500000.00, 45000.00, 1850000.00, 9.2, 60, 42
            ]
        ]

    def get_sample_data_leasing(self):
        """Sample data for leasing template"""
        return [
            [
                "CENTRAL", "LS001234567", "Carol Brown", "567890123V", "654 Fifth St, Negombo",
                "0773333333", "0312345678", "", 3500000.00, "2022-09-01", "2025-09-01",
                85000.00, "2025-07-01", 80000.00, "2025-08-01", "Sales Manager", "JKL Motors",
                "666 Auto St, Negombo", "0312223344", "0319876543", "", "Tom Brown",
                "777 Father St, Katunayake", "890123456V", "0777777777", "0312121212", "Emma Davis",
                "888 Mother St, Seeduwa", "901234567V", "0778888888", "", "Western", "Gampaha",
                "Negombo", "Gampaha", "Negombo", 18, "Vehicle lease customer", "Monthly rentals",
                "CAR-1234", "Toyota", "Prius", "2022", 3500000.00, 36, 85000.00, 2100000.00, 800000.00
            ],
            [
                "SANASA", "LS987654321", "David Lee", "678901234V", "987 Sixth St, Anuradhapura",
                "0774444444", "0252345678", "0779999999", 2800000.00, "2023-02-15", "2026-02-15",
                65000.00, "2025-06-15", 60000.00, "2025-08-15", "Driver", "MNO Transport",
                "999 Transport St, Anuradhapura", "0252223344", "", "", "Mary Lee",
                "111 Spouse St, Mihintale", "012345678V", "0770000000", "", "", "", "", "", "", "",
                "North Central", "Anuradhapura", "Anuradhapura", "Anuradhapura", "Anuradhapura", 3,
                "Van lease for business", "Good payment record", "VAN-5678", "Mitsubishi", "L300",
                "2023", 2800000.00, 36, 65000.00, 1950000.00, 500000.00
            ]
        ]

    def apply_header_styling(self, ws, headers):
        """Apply consistent header styling"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header.replace("_", " ").title()
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

    def auto_adjust_columns(self, ws, headers, sample_data=None):
        """Auto-adjust column widths"""
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0

            # Check header length
            header_cell = ws.cell(row=1, column=col_num)
            if header_cell.value:
                max_length = max(max_length, len(str(header_cell.value)))

            # Check sample data length if exists
            if sample_data:
                for row_num in range(2, len(sample_data) + 2):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

            # Set width with reasonable limits
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def add_title_and_instructions(self, ws, title, operation_type):
        """Add title and instructions to the template"""
        # Insert rows at the top
        ws.insert_rows(1, 4)

        # Add title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(name="Arial", size=16, bold=True, color="366092")
        title_cell.alignment = Alignment(horizontal="center")

        # Add instructions
        ws.merge_cells('A2:F2')
        instructions_cell = ws['A2']
        instructions_cell.value = f"Import Template for {operation_type.replace('_', ' ').title()} Operations"
        instructions_cell.font = Font(name="Arial", size=12, color="666666")
        instructions_cell.alignment = Alignment(horizontal="center")

        # Add important notes
        ws.merge_cells('A3:F3')
        notes_cell = ws['A3']
        notes_cell.value = "Important: client_code must match the selected client. Fill all required fields marked with *"
        notes_cell.font = Font(name="Arial", size=10, color="CC0000")
        notes_cell.alignment = Alignment(horizontal="center")

        # Headers will now be in row 5
        return 5

    def generate_template(self, operation_type: str, include_sample: bool = True):
        """Generate comprehensive Excel template"""

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{operation_type}_Import_Template"

        # Get base customer fields
        base_headers = self.get_all_customer_fields()

        # Add operation-specific fields
        if operation_type == "CREDIT_CARD":
            headers = base_headers + self.get_credit_card_specific_fields()
            sample_data = self.get_sample_data_credit_card() if include_sample else None
            title = "RECOVA - Credit Card Import Template"

        elif operation_type == "LOAN":
            headers = base_headers + self.get_loan_specific_fields()
            sample_data = self.get_sample_data_loan() if include_sample else None
            title = "RECOVA - Loan Import Template"

        elif operation_type == "LEASING":
            headers = base_headers + self.get_leasing_specific_fields()
            sample_data = self.get_sample_data_leasing() if include_sample else None
            title = "RECOVA - Leasing Import Template"

        else:
            headers = base_headers
            sample_data = None
            title = f"RECOVA - {operation_type} Import Template"

        # Add title and instructions
        header_row = self.add_title_and_instructions(ws, title, operation_type)

        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header.replace("_", " ").title()

            # Mark required fields
            if header in ["client_code", "contract_number", "client_name", "nic"]:
                cell.value += " *"

            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

        # Add sample data if requested
        if include_sample and sample_data:
            for row_idx, row_data in enumerate(sample_data, header_row + 1):
                for col_idx, value in enumerate(row_data, 1):
                    if col_idx <= len(headers):  # Don't exceed header count
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.border = self.thin_border

        # Auto-adjust columns
        self.auto_adjust_columns(ws, headers, sample_data)

        # Freeze the header row
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Save template
        template_dir = "templates"
        os.makedirs(template_dir, exist_ok=True)
        filename = f"RECOVA_{operation_type}_Template.xlsx"
        filepath = os.path.join(template_dir, filename)
        wb.save(filepath)

        return filepath

    def get_template_info(self):
        """Get template information for API responses"""
        return {
            "CREDIT_CARD": {
                "description": "Template for importing credit card defaulter accounts with complete customer information",
                "field_count": len(self.get_all_customer_fields()) + len(self.get_credit_card_specific_fields()),
                "sample_provided": True,
                "required_fields": ["client_code", "contract_number", "client_name", "nic"]
            },
            "LOAN": {
                "description": "Template for importing loan defaulter accounts with complete customer information",
                "field_count": len(self.get_all_customer_fields()) + len(self.get_loan_specific_fields()),
                "sample_provided": True,
                "required_fields": ["client_code", "contract_number", "client_name", "nic"]
            },
            "LEASING": {
                "description": "Template for importing leasing/vehicle finance defaulter accounts with complete customer information",
                "field_count": len(self.get_all_customer_fields()) + len(self.get_leasing_specific_fields()),
                "sample_provided": True,
                "required_fields": ["client_code", "contract_number", "client_name", "nic"]
            }
        }