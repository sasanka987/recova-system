# File: app/services/excel_processor.py
import openpyxl
import pandas as pd
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime, date
import re
from decimal import Decimal, InvalidOperation


class ExcelProcessor:
    """Handle Excel file processing and validation"""

    def __init__(self):
        self.required_columns = {
            "CREDIT_CARD": [
                "client_name", "nic", "card_number", "account_number",
                "credit_limit", "total_outstanding_amount", "minimum_due_amount",
                "contract_number"
            ],
            "LOAN": [
                "client_name", "nic", "loan_number", "account_number",
                "granted_amount", "capital_balance", "interest_over_due_balance",
                "contract_number"
            ],
            "LEASING": [
                "client_name", "nic", "vehicle_number", "asset_description",
                "total_arrears", "rentals_in_arrears", "contract_number"
            ],
            "PAYMENT": [
                "payment_date", "contract_number", "payment_amount",
                "payment_type", "receipt_number"
            ]
        }

    def read_excel_file(self, file_path: str, operation_type: str) -> Tuple[List[Dict], List[Dict]]:
        """
        Read Excel file and return data with validation errors
        Returns: (valid_records, validation_errors)
        """
        try:
            # Read Excel file
            df = pd.read_excel(file_path, engine='openpyxl')

            # Clean column names
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

            # Validate required columns
            required_cols = self.required_columns.get(operation_type, [])
            missing_columns = [col for col in required_cols if col not in df.columns]

            if missing_columns:
                raise ValueError(f"Missing required columns: {missing_columns}")

            # Process rows
            valid_records = []
            validation_errors = []

            for index, row in df.iterrows():
                try:
                    # Clean and validate row data
                    cleaned_row = self._clean_row_data(row, operation_type)
                    validation_result = self._validate_row(cleaned_row, operation_type,
                                                           index + 2)  # +2 for Excel row number

                    if validation_result["is_valid"]:
                        valid_records.append(cleaned_row)
                    else:
                        validation_errors.extend(validation_result["errors"])

                except Exception as e:
                    validation_errors.append({
                        "row_number": index + 2,
                        "column_name": None,
                        "error_type": "ROW_PROCESSING_ERROR",
                        "error_message": str(e),
                        "original_value": str(row.to_dict()),
                        "is_critical": True
                    })

            return valid_records, validation_errors

        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")

    def _clean_row_data(self, row: pd.Series, operation_type: str) -> Dict[str, Any]:
        """Clean and standardize row data"""
        cleaned = {}

        for column, value in row.items():
            # Handle NaN values
            if pd.isna(value):
                cleaned[column] = None
                continue

            # Clean string values
            if isinstance(value, str):
                cleaned[column] = value.strip()
            else:
                cleaned[column] = value

        return cleaned

    def _validate_row(self, row_data: Dict, operation_type: str, row_number: int) -> Dict:
        """Validate individual row data"""
        errors = []
        is_valid = True

        # Validate required fields
        required_cols = self.required_columns.get(operation_type, [])
        for col in required_cols:
            if not row_data.get(col):
                errors.append({
                    "row_number": row_number,
                    "column_name": col,
                    "error_type": "REQUIRED_FIELD_MISSING",
                    "error_message": f"Required field '{col}' is missing or empty",
                    "original_value": row_data.get(col),
                    "is_critical": True
                })
                is_valid = False

        # Validate NIC format
        if row_data.get("nic"):
            if not self._validate_nic(row_data["nic"]):
                errors.append({
                    "row_number": row_number,
                    "column_name": "nic",
                    "error_type": "INVALID_NIC_FORMAT",
                    "error_message": "Invalid NIC format",
                    "original_value": row_data["nic"],
                    "suggested_value": "Use format: 123456789V or 123456789012",
                    "is_critical": True
                })
                is_valid = False

        # Validate phone numbers
        phone_fields = ["customer_contact_number_1", "customer_contact_number_2", "customer_contact_number_3"]
        for field in phone_fields:
            if row_data.get(field) and not self._validate_phone_number(row_data[field]):
                errors.append({
                    "row_number": row_number,
                    "column_name": field,
                    "error_type": "INVALID_PHONE_FORMAT",
                    "error_message": "Invalid phone number format",
                    "original_value": row_data[field],
                    "suggested_value": "Use format: 0771234567",
                    "is_critical": False
                })

        # Validate monetary amounts
        amount_fields = ["payment_amount", "granted_amount", "credit_limit", "total_outstanding_amount"]
        for field in amount_fields:
            if row_data.get(field) is not None:
                try:
                    amount = Decimal(str(row_data[field]))
                    if amount < 0:
                        errors.append({
                            "row_number": row_number,
                            "column_name": field,
                            "error_type": "NEGATIVE_AMOUNT",
                            "error_message": f"Amount cannot be negative",
                            "original_value": row_data[field],
                            "is_critical": True
                        })
                        is_valid = False
                except (InvalidOperation, ValueError):
                    errors.append({
                        "row_number": row_number,
                        "column_name": field,
                        "error_type": "INVALID_AMOUNT_FORMAT",
                        "error_message": f"Invalid amount format",
                        "original_value": row_data[field],
                        "is_critical": True
                    })
                    is_valid = False

        # Validate dates
        date_fields = ["payment_date", "facility_granted_date", "facility_end_date", "due_date"]
        for field in date_fields:
            if row_data.get(field) and not self._validate_date(row_data[field]):
                errors.append({
                    "row_number": row_number,
                    "column_name": field,
                    "error_type": "INVALID_DATE_FORMAT",
                    "error_message": "Invalid date format",
                    "original_value": row_data[field],
                    "suggested_value": "Use format: YYYY-MM-DD",
                    "is_critical": True
                })
                is_valid = False

        return {
            "is_valid": is_valid,
            "errors": errors
        }

    def _validate_nic(self, nic: str) -> bool:
        """Validate Sri Lankan NIC format"""
        if not nic:
            return False

        # Old format: 9 digits + V
        old_format = re.match(r'^\d{9}[VvXx]$', nic)
        # New format: 12 digits
        new_format = re.match(r'^\d{12}$', nic)

        return bool(old_format or new_format)

    def _validate_phone_number(self, phone: str) -> bool:
        """Validate Sri Lankan phone number format"""
        if not phone:
            return True  # Optional field

        # Remove spaces and special characters
        clean_phone = re.sub(r'[^\d]', '', phone)

        # Sri Lankan mobile: 07XXXXXXXX (10 digits)
        # Landline: 0XXXXXXX (9-10 digits)
        return len(clean_phone) >= 9 and len(clean_phone) <= 10 and clean_phone.startswith('0')

    def _validate_date(self, date_value) -> bool:
        """Validate date format"""
        if not date_value:
            return True  # Optional field

        try:
            if isinstance(date_value, (date, datetime)):
                return True

            # Try parsing string dates
            pd.to_datetime(date_value)
            return True
        except:
            return False


# File: app/services/import_service.py
from sqlalchemy.orm import Session
from typing import List, Dict, Any
import os
from datetime import datetime
import hashlib

from app.models.import_batch import ImportBatch, ImportError, ImportStatus, OperationType
from app.models.customer import Customer
from app.models.payment import Payment
from app.schemas.import_batch import ImportBatchCreate, ImportSummary
from app.services.excel_processor import ExcelProcessor


class ImportService:
    """Main service for handling data imports"""

    def __init__(self, db: Session):
        self.db = db
        self.excel_processor = ExcelProcessor()

    def create_import_batch(self, batch_data: ImportBatchCreate, file_path: str, file_size: int,
                            user_id: int) -> ImportBatch:
        """Create a new import batch record"""

        # Extract month and year from import_period
        # e.g., "August 2025" -> month=8, year=2025
        try:
            month_name, year_str = batch_data.import_period.split()
            month_map = {
                "January": 1, "February": 2, "March": 3, "April": 4,
                "May": 5, "June": 6, "July": 7, "August": 8,
                "September": 9, "October": 10, "November": 11, "December": 12
            }
            month = month_map.get(month_name, 1)
            year = int(year_str)
        except:
            month, year = datetime.now().month, datetime.now().year

        # Calculate file checksum
        checksum = self._calculate_file_checksum(file_path)

        # Create batch record
        batch = ImportBatch(
            batch_name=batch_data.batch_name,
            bank_name=batch_data.bank_name,
            bank_code=batch_data.bank_code,
            operation_type=OperationType(batch_data.operation_type),
            import_period=batch_data.import_period,
            import_month=month,
            import_year=year,
            file_name=batch_data.file_name,
            file_size=file_size,
            file_path=file_path,
            file_checksum=checksum,
            status=ImportStatus.UPLOADED,
            imported_by=user_id
        )

        self.db.add(batch)
        self.db.commit()
        self.db.refresh(batch)

        return batch

    def validate_import_batch(self, batch_id: int) -> ImportSummary:
        """Validate the data in an import batch"""

        batch = self.db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
        if not batch:
            raise ValueError("Import batch not found")

        # Update status to validating
        batch.status = ImportStatus.VALIDATING
        self.db.commit()

        try:
            # Process Excel file
            valid_records, validation_errors = self.excel_processor.read_excel_file(
                batch.file_path,
                batch.operation_type.value
            )

            # Update batch statistics
            batch.total_records = len(valid_records) + len(validation_errors)
            batch.valid_records = len(valid_records)
            batch.invalid_records = len(validation_errors)

            # Save validation errors
            for error_data in validation_errors:
                error = ImportError(
                    batch_id=batch_id,
                    row_number=error_data["row_number"],
                    column_name=error_data.get("column_name"),
                    error_type=error_data["error_type"],
                    error_message=error_data["error_message"],
                    original_value=error_data.get("original_value"),
                    suggested_value=error_data.get("suggested_value"),
                    is_critical=error_data["is_critical"]
                )
                self.db.add(error)

            # Update batch status
            if len(validation_errors) == 0:
                batch.status = ImportStatus.VALIDATED
            elif any(error["is_critical"] for error in validation_errors):
                batch.status = ImportStatus.FAILED
            else:
                batch.status = ImportStatus.VALIDATED  # Non-critical errors

            self.db.commit()

            # Return validation summary
            return ImportSummary(
                batch_id=batch_id,
                batch_name=batch.batch_name,
                status=batch.status,
                total_records=batch.total_records,
                valid_records=batch.valid_records,
                invalid_records=batch.invalid_records,
                imported_records=0,
                errors=[
                    {
                        "id": 0,
                        "row_number": err["row_number"],
                        "column_name": err.get("column_name"),
                        "error_type": err["error_type"],
                        "error_message": err["error_message"],
                        "original_value": err.get("original_value"),
                        "suggested_value": err.get("suggested_value"),
                        "is_critical": err["is_critical"]
                    }
                    for err in validation_errors
                ]
            )

        except Exception as e:
            batch.status = ImportStatus.FAILED
            self.db.commit()
            raise Exception(f"Validation failed: {str(e)}")

    def process_import_batch(self, batch_id: int) -> ImportSummary:
        """Process validated data and import into database"""

        batch = self.db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
        if not batch:
            raise ValueError("Import batch not found")

        if batch.status != ImportStatus.VALIDATED:
            raise ValueError("Batch must be validated before processing")

        # Update status to importing
        batch.status = ImportStatus.IMPORTING
        batch.import_started_at = datetime.now()
        self.db.commit()

        try:
            # Read valid records from Excel
            valid_records, _ = self.excel_processor.read_excel_file(
                batch.file_path,
                batch.operation_type.value
            )

            imported_count = 0

            if batch.operation_type == OperationType.PAYMENT:
                # Process payment records
                imported_count = self._process_payment_records(valid_records, batch_id)
            else:
                # Process customer/contract records
                imported_count = self._process_customer_records(valid_records, batch_id, batch.operation_type)

            # Update batch completion
            batch.status = ImportStatus.COMPLETED
            batch.imported_records = imported_count
            batch.import_completed_at = datetime.now()
            self.db.commit()

            return ImportSummary(
                batch_id=batch_id,
                batch_name=batch.batch_name,
                status=batch.status,
                total_records=batch.total_records,
                valid_records=batch.valid_records,
                invalid_records=batch.invalid_records,
                imported_records=imported_count,
                errors=[],
                processing_time_seconds=(batch.import_completed_at - batch.import_started_at).total_seconds()
            )

        except Exception as e:
            batch.status = ImportStatus.FAILED
            self.db.commit()
            raise Exception(f"Import processing failed: {str(e)}")

    def _process_customer_records(self, records: List[Dict], batch_id: int, operation_type: OperationType) -> int:
        """Process customer and contract records"""
        imported_count = 0

        for record in records:
            try:
                # Check if customer already exists
                existing_customer = self.db.query(Customer).filter(
                    Customer.nic == record.get("nic")
                ).first()

                if existing_customer:
                    # Update existing customer
                    for key, value in record.items():
                        if hasattr(existing_customer, key) and value is not None:
                            setattr(existing_customer, key, value)
                    existing_customer.import_batch_id = batch_id
                else:
                    # Create new customer
                    customer_data = {k: v for k, v in record.items() if hasattr(Customer, k)}
                    customer_data["import_batch_id"] = batch_id
                    customer = Customer(**customer_data)
                    self.db.add(customer)

                imported_count += 1

                # Commit every 100 records for performance
                if imported_count % 100 == 0:
                    self.db.commit()

            except Exception as e:
                print(f"Error processing record: {e}")
                continue

        self.db.commit()
        return imported_count

    def _process_payment_records(self, records: List[Dict], batch_id: int) -> int:
        """Process payment records with automatic matching"""
        imported_count = 0

        for record in records:
            try:
                # Create payment record
                payment_data = {k: v for k, v in record.items() if hasattr(Payment, k)}
                payment_data["import_batch_id"] = batch_id

                # Attempt to match payment to customer
                match_result = self._match_payment_to_customer(record)
                payment_data["matched_customer_id"] = match_result["customer_id"]
                payment_data["match_type"] = match_result["match_type"]
                payment_data["is_matched"] = match_result["customer_id"] is not None

                payment = Payment(**payment_data)
                self.db.add(payment)

                # Update customer balance if matched
                if payment_data["is_matched"]:
                    self._update_customer_balance(payment_data["matched_customer_id"], record)

                imported_count += 1

                # Commit every 100 records
                if imported_count % 100 == 0:
                    self.db.commit()

            except Exception as e:
                print(f"Error processing payment record: {e}")
                continue

        self.db.commit()
        return imported_count

    def _match_payment_to_customer(self, payment_record: Dict) -> Dict:
        """Match payment to existing customer using contract number, account number, or NIC"""

        # Try contract number first
        if payment_record.get("contract_number"):
            customer = self.db.query(Customer).filter(
                Customer.contract_number == payment_record["contract_number"]
            ).first()
            if customer:
                return {"customer_id": customer.id, "match_type": "CONTRACT_NUMBER_MATCH"}

        # Try account number (need to check contract tables)
        if payment_record.get("account_number"):
            # This would require checking CreditCardContract, LoanAdvanceContract, etc.
            # For now, simplified approach
            pass

        # Try customer NIC
        if payment_record.get("customer_nic"):
            customer = self.db.query(Customer).filter(
                Customer.nic == payment_record["customer_nic"]
            ).first()
            if customer:
                return {"customer_id": customer.id, "match_type": "NIC_MATCH"}

        return {"customer_id": None, "match_type": "NO_MATCH"}

    def _update_customer_balance(self, customer_id: int, payment_record: Dict):
        """Update customer balance after payment"""
        customer = self.db.query(Customer).filter(Customer.id == customer_id).first()
        if customer and payment_record.get("current_total_arrears"):
            # Update last payment details
            customer.last_payment_date = payment_record.get("payment_date")
            customer.last_payment_amount = payment_record.get("payment_amount")

            # Calculate days in arrears (simplified)
            if customer.due_date and payment_record.get("payment_date"):
                days_diff = (payment_record["payment_date"] - customer.due_date).days
                customer.days_in_arrears = max(0, days_diff)

    def get_import_batch(self, batch_id: int) -> ImportBatch:
        """Get import batch by ID"""
        return self.db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()

    def get_import_errors(self, batch_id: int) -> List[ImportError]:
        """Get validation errors for a batch"""
        return self.db.query(ImportError).filter(ImportError.batch_id == batch_id).all()

    def get_import_batches(self, skip: int = 0, limit: int = 100) -> List[ImportBatch]:
        """Get list of import batches"""
        return self.db.query(ImportBatch).offset(skip).limit(limit).all()

    def _calculate_file_checksum(self, file_path: str) -> str:
        """Calculate MD5 checksum of file"""
        hash_md5 = hashlib.md5()
        try:
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except:
            return ""


# File: app/services/template_service.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, List
import os


class TemplateService:
    """Service for generating Excel templates"""

    def __init__(self):
        self.templates = {
            "CREDIT_CARD": {
                "headers": [
                    "client_name", "nic", "home_address", "customer_contact_number_1",
                    "customer_contact_number_2", "customer_contact_number_3", "contract_number",
                    "card_product_type", "card_number", "account_number", "credit_limit",
                    "capital_balance", "interest_over_due_balance", "total_outstanding_amount",
                    "minimum_due_amount", "statement_date", "payment_due_date",
                    "work_place_name", "work_place_address", "zone", "region", "branch_name",
                    "guarantor_1_name", "guarantor_1_nic", "guarantor_1_contact_number_1"
                ],
                "sample_data": [
                    ["John Doe", "123456789V", "123 Main St, Colombo", "0771234567",
                     "0112345678", "", "CC001234567", "VISA GOLD", "1234567890123456",
                     "ACC123456", 500000, 250000, 15000, 265000, 25000, "2025-01-15",
                     "2025-02-10", "ABC Company", "456 Business Ave", "Colombo",
                     "Western", "Colombo Central", "Jane Doe", "987654321V", "0779876543"]
                ]
            },
            "LOAN": {
                "headers": [
                    "client_name", "nic", "home_address", "customer_contact_number_1",
                    "contract_number", "loan_type", "granted_amount", "capital_balance",
                    "interest_over_due_balance", "installment_amount", "operative_account_number",
                    "security_description", "security_value", "work_place_name",
                    "guarantor_1_name", "guarantor_1_nic"
                ],
                "sample_data": [
                    ["Alice Smith", "234567890V", "789 Oak Road, Kandy", "0712345678",
                     "LN987654321", "PERSONAL", 1000000, 750000, 50000, 15000,
                     "OP123456789", "Property at Kandy", 1500000, "XYZ Corporation",
                     "Bob Smith", "345678901V"]
                ]
            },
            "LEASING": {
                "headers": [
                    "client_name", "nic", "home_address", "customer_contact_number_1",
                    "contract_number", "vehicle_number", "asset_description", "model",
                    "total_arrears", "rentals_in_arrears", "rental_category",
                    "other_party_name", "other_party_contact_number"
                ],
                "sample_data": [
                    ["Mike Johnson", "345678901V", "321 Pine St, Galle", "0723456789",
                     "LS456789012", "CAR-1234", "Toyota Corolla 2020", "Corolla",
                     150000, 3, "Vehicle Leasing", "Sarah Johnson", "0734567890"]
                ]
            },
            "PAYMENT": {
                "headers": [
                    "payment_date", "contract_number", "account_number", "customer_nic",
                    "payment_amount", "payment_type", "receipt_number", "bank_reference_number",
                    "payment_method", "branch_name"
                ],
                "sample_data": [
                    ["2025-08-13", "CC001234567", "ACC123456", "123456789V",
                     25000, "PARTIAL_PAYMENT", "RCP001", "BNK123456", "Bank Transfer",
                     "Colombo Central"]
                ]
            }
        }

    def generate_template(self, operation_type: str, include_sample: bool = True) -> str:
        """Generate Excel template for given operation type"""

        if operation_type not in self.templates:
            raise ValueError(f"Template not available for operation type: {operation_type}")

        template_data = self.templates[operation_type]

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{operation_type}_Template"

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers
        for col, header in enumerate(template_data["headers"], 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header.replace("_", " ").title()
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add sample data if requested
        if include_sample and template_data.get("sample_data"):
            for row_idx, sample_row in enumerate(template_data["sample_data"], 2):
                for col_idx, value in enumerate(sample_row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save template
        template_dir = "templates"
        os.makedirs(template_dir, exist_ok=True)
        filename = f"{operation_type.lower()}_template.xlsx"
        filepath = os.path.join(template_dir, filename)
        wb.save(filepath)

        return filepath